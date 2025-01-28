import pandas as pd
from openpyxl import load_workbook
from pymongo import MongoClient
import uuid

def read_excel(file_path):
    workbook = load_workbook(file_path, data_only=True)
    sheets = workbook.sheetnames
    return workbook, sheets

def parse_sheet(workbook, sheet_name):
    sheet = workbook[sheet_name]
    data = []
    headers = []

    for row in sheet.iter_rows():
        if any(cell.value for cell in row): 
            headers = [str(cell.value) if cell.value else "Unnamed Column" for cell in row]
            break

    for row in sheet.iter_rows(min_row=sheet.min_row + 1, values_only=True):
        record = {str(headers[i]): row[i] for i in range(len(headers))}
        data.append(record)

    return data

def assign_uuids(data):
    for record in data:
        record["_id"] = str(uuid.uuid4())
    return data

def insert_into_mongo(data, db_name, collection_name):
    client = MongoClient("mongodb://localhost:27017/")
    db = client[db_name]
    collection = db["collection_name"]
    collection.insert_many(data)

    print(f"Inserted {len(data)} records into {db_name}.{collection_name}")

def process_excel_to_mongo(file_path, db_name):
    workbook, sheets = read_excel(file_path)
    for sheet_name in sheets:
        print(f"Parsing sheet: {sheet_name}")
        data = parse_sheet(workbook, sheet_name)
        data = assign_uuids(data)
        insert_into_mongo(data, db_name, sheet_name)

file_path = "samp.xlsx"
db_name = "excel_data"
process_excel_to_mongo(file_path, db_name)